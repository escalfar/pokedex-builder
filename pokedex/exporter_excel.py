from __future__ import annotations

from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pokedex.constants import OUTPUT_COLUMNS, ExcelSheet, GameColumn, OutputColumn
from pokedex.exceptions import PokedexError
from pokedex.models import PokemonEntry

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=16, bold=True)
_SUBTITLE_FONT = Font(size=11, italic=True)
_BOOLEAN_TRUE_FILL = PatternFill("solid", fgColor="E2F0D9")
_BOOLEAN_FALSE_FILL = PatternFill("solid", fgColor="FCE4D6")

_EXCEL_BASE_FORM_LABELS: dict[str, str] = {
    "Deoxys": "Normal",
    "Wormadam": "Plant",
    "Shaymin": "Land",
    "Basculin": "Red",
    "Pumpkaboo": "Medium",
    "Gourgeist": "Medium",
    "Oricorio": "Baile",
    "Lycanroc": "Midday",
    "Toxtricity": "Amped",
    "Urshifu": "Single",
    "Squawkabilly": "Green",
    "Tatsugiri": "Curly",
    "Gimmighoul": "Chest",
}

_EXCEL_FORM_LABEL_OVERRIDES: dict[str, dict[str, str]] = {
    "Tauros": {
        "Paldean Aqua Breed": "Paldean Aqua",
        "Paldean Blaze Breed": "Paldean Blaze",
        "Paldean Combat Breed": "Paldean Combat",
    },
    "Burmy": {
        "Plant Cloak": "Plant",
        "Sandy Cloak": "Sandy",
        "Trash Cloak": "Trash",
    },
    "Basculin": {
        "Blue Striped": "Blue",
        "White Striped": "White",
    },
    "Pumpkaboo": {"Super": "Jumbo"},
    "Gourgeist": {"Super": "Jumbo"},
    "Urshifu": {"Rapid Strike": "Rapid"},
    "Squawkabilly": {
        "Blue Plumage": "Blue",
        "White Plumage": "White",
        "Yellow Plumage": "Yellow",
    },
}

_FLOWER_FORM_ORDER: dict[str, int] = {
    "Red Flower": 0,
    "Yellow Flower": 1,
    "Orange Flower": 2,
    "Blue Flower": 3,
    "White Flower": 4,
    "Eternal": 5,
}
_FLOWER_SPECIES = {"Flabébé", "Floette", "Florges"}


def export_excel(
    entries: tuple[PokemonEntry, ...],
    output_path: Path,
    *,
    generator_version: str = "0.1.0",
    specification_version: str = "1.0",
    generated_at: datetime | None = None,
) -> Path:
    """Export entries to a formatted workbook with four worksheets.

    The workbook intentionally contains values rather than formulas so it can
    be consumed consistently by Excel, LibreOffice, Power BI, and pandas.
    """
    if not entries:
        raise ValueError("Cannot export an empty Pokémon entry collection.")

    timestamp = generated_at or datetime.now(UTC)
    workbook = Workbook()

    try:
        pokedex_sheet = workbook.active
        # ``Workbook.active`` may be typed as a worksheet, chartsheet, or None.
        # The exporter requires a real worksheet before populating cells.
        if not isinstance(pokedex_sheet, Worksheet):
            raise PokedexError("Workbook does not contain an active worksheet.")

        pokedex_sheet.title = ExcelSheet.POKEDEX.value
        _populate_pokedex_sheet(pokedex_sheet, entries)

        summary_sheet = workbook.create_sheet(ExcelSheet.SUMMARY.value)
        _populate_summary_sheet(summary_sheet, entries)

        validation_sheet = workbook.create_sheet(ExcelSheet.VALIDATION.value)
        _populate_validation_sheet(validation_sheet, entries)

        metadata_sheet = workbook.create_sheet(ExcelSheet.METADATA.value)
        _populate_metadata_sheet(
            metadata_sheet,
            entries,
            generator_version=generator_version,
            specification_version=specification_version,
            generated_at=timestamp,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    except (OSError, ValueError, TypeError) as error:
        _remove_temporary_file(locals().get("temporary_path"))
        raise PokedexError(f"Unable to export Excel file: {output_path}") from error
    finally:
        workbook.close()

    return output_path


def _populate_pokedex_sheet(
    sheet: Worksheet, entries: tuple[PokemonEntry, ...]
) -> None:
    # openpyxl worksheets are duck-typed here to avoid depending on private
    # worksheet implementation types in the public function signature.
    headers = [column.value for column in OUTPUT_COLUMNS]
    sheet.append(headers)

    for entry in _excel_ordered_entries(entries):
        export_row = entry.to_dict()
        # The Excel presentation groups every form under its species name.
        # CSV, JSON, and all domain objects retain the original internal name.
        export_row["Nombre"] = _excel_display_name(entry)
        sheet.append([export_row[header] for header in headers])

    sheet.freeze_panes = "A2"
    # Use a worksheet-level filter instead of an OOXML structured table.
    # This preserves filtering in Excel while avoiding table repair warnings.
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    boolean_columns = {
        index + 1
        for index, header in enumerate(headers)
        if header in {game.value for game in GameColumn}
        or header in {"Legendario/Mítico", "Obtenible"}
    }

    for worksheet_row in sheet.iter_rows(min_row=2):
        for cell in worksheet_row:
            if cell.column in boolean_columns:
                cell.alignment = Alignment(horizontal="center")
                cell.fill = (
                    _BOOLEAN_TRUE_FILL if cell.value is True else _BOOLEAN_FALSE_FILL
                )

    _set_column_widths(sheet, headers)
    _hide_internal_columns(sheet, headers)


def _hide_internal_columns(sheet: Worksheet, headers: list[str]) -> None:
    """Hide internal and reference columns in the primary Excel worksheet."""
    hidden_headers = {
        OutputColumn.POKEMON.value,
        OutputColumn.FORM.value,
        OutputColumn.GENERATION.value,
        OutputColumn.HOME_ID.value,
        OutputColumn.LEGENDARY_MYTHICAL.value,
        OutputColumn.OBTAINABLE_SHINY.value,
    }

    for index, header in enumerate(headers, start=1):
        if header in hidden_headers:
            sheet.column_dimensions[get_column_letter(index)].hidden = True


def _excel_ordered_entries(
    entries: tuple[PokemonEntry, ...],
) -> tuple[PokemonEntry, ...]:
    """Return the Excel-only order without changing the domain collection."""
    indexed_entries = tuple(enumerate(entries))

    def sort_key(item: tuple[int, PokemonEntry]) -> tuple[int, int]:
        original_index, entry = item
        if entry.pokemon in _FLOWER_SPECIES:
            return (entry.national_dex, _FLOWER_FORM_ORDER.get(entry.form, 99))
        return (entry.national_dex, original_index)

    return tuple(entry for _, entry in sorted(indexed_entries, key=sort_key))


def _excel_display_name(entry: PokemonEntry) -> str:
    """Return the Excel-only display name without changing domain data."""
    if entry.form.casefold() == "normal":
        base_form = _EXCEL_BASE_FORM_LABELS.get(entry.pokemon)
        return f"{entry.pokemon} ({base_form})" if base_form else entry.pokemon

    display_form = _excel_display_form(entry.pokemon, entry.form)
    return f"{entry.pokemon} ({display_form})"


def _excel_display_form(pokemon: str, form: str) -> str:
    """Normalize selected form labels exclusively for the Excel Nombre column."""
    override = _EXCEL_FORM_LABEL_OVERRIDES.get(pokemon, {}).get(form)
    if override is not None:
        return override

    if pokemon == "Vivillon" and form.endswith(" Pattern"):
        return form.removesuffix(" Pattern")

    if pokemon in _FLOWER_SPECIES and form.endswith(" Flower"):
        return form.removesuffix(" Flower")

    return form


def _populate_summary_sheet(
    sheet: Worksheet, entries: tuple[PokemonEntry, ...]
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Resumen de la Pokédex"
    sheet["A1"].font = _TITLE_FONT
    sheet["A3"] = "Métrica"
    sheet["B3"] = "Valor"

    unique_species = len({entry.national_dex for entry in entries})
    legendary_count = sum(entry.legendary_mythical for entry in entries)
    shiny_count = sum(entry.obtainable_shiny for entry in entries)

    metrics: list[tuple[str, int]] = [
        ("Total de filas", len(entries)),
        ("Total de especies", unique_species),
        ("Total de variantes adicionales", len(entries) - unique_species),
        ("Legendarios/Míticos", legendary_count),
        ("Shiny obtenible sin evento", shiny_count),
    ]

    for game in GameColumn:
        metrics.append(
            (
                f"Obtenibles en {game.value}",
                sum(entry.availability.is_available_in(game) for entry in entries),
            )
        )

    for row_number, (label, value) in enumerate(metrics, start=4):
        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value)

    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 16


def _populate_validation_sheet(
    sheet: Worksheet, entries: tuple[PokemonEntry, ...]
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Validación"
    sheet["A1"].font = _TITLE_FONT
    sheet["A3"] = "Comprobación"
    sheet["B3"] = "Estado"
    sheet["C3"] = "Detalle"

    checks = _validation_checks(entries)
    for row_number, (name, valid, detail) in enumerate(checks, start=4):
        sheet.cell(row=row_number, column=1, value=name)
        status_cell = sheet.cell(
            row=row_number,
            column=2,
            value="OK" if valid else "ERROR",
        )
        status_cell.fill = _BOOLEAN_TRUE_FILL if valid else _BOOLEAN_FALSE_FILL
        sheet.cell(row=row_number, column=3, value=detail)

    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 70


def _populate_metadata_sheet(
    sheet: Worksheet,
    entries: tuple[PokemonEntry, ...],
    *,
    generator_version: str,
    specification_version: str,
    generated_at: datetime,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Metadatos"
    sheet["A1"].font = _TITLE_FONT
    sheet["A2"] = "Información de generación del conjunto de datos"
    sheet["A2"].font = _SUBTITLE_FONT

    values: tuple[tuple[str, str | int], ...] = (
        ("Generator Version", generator_version),
        ("Specification", specification_version),
        ("Generation Date (UTC)", generated_at.astimezone(UTC).isoformat()),
        ("Total Records", len(entries)),
        ("National Dex Max", max(entry.national_dex for entry in entries)),
        ("Output Columns", len(OUTPUT_COLUMNS)),
    )

    for row_number, (label, value) in enumerate(values, start=4):
        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value)

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 40


def _validation_checks(
    entries: tuple[PokemonEntry, ...],
) -> tuple[tuple[str, bool, str], ...]:
    home_id_counts = Counter(entry.home_id for entry in entries)
    duplicate_home_ids = sorted(
        home_id for home_id, count in home_id_counts.items() if count > 1
    )

    name_counts = Counter(entry.name.casefold() for entry in entries)
    duplicate_names = sorted(name for name, count in name_counts.items() if count > 1)

    expected_columns = {column.value for column in OUTPUT_COLUMNS}
    invalid_rows = [
        entry.home_id for entry in entries if set(entry.to_dict()) != expected_columns
    ]

    return (
        (
            "ID HOME único",
            not duplicate_home_ids,
            ", ".join(duplicate_home_ids) or "Sin duplicados",
        ),
        (
            "Nombre único",
            not duplicate_names,
            ", ".join(duplicate_names) or "Sin duplicados",
        ),
        (
            "Columnas completas",
            not invalid_rows,
            ", ".join(invalid_rows) or "Todas las filas tienen 19 columnas",
        ),
        (
            "Orden por Pokédex",
            tuple(entry.national_dex for entry in entries)
            == tuple(sorted(entry.national_dex for entry in entries)),
            "Orden ascendente verificado",
        ),
    )


def _set_column_widths(sheet: Worksheet, headers: list[str]) -> None:
    preferred_widths = {
        "Nat Dex": 10,
        "Pokemon": 20,
        "Forma": 24,
        "Nombre": 34,
        "Obtenido": 14,
        "Gen": 8,
        "ID HOME": 30,
        "Legendario/Mítico": 20,
        "Obtenible": 14,
    }

    for index, header in enumerate(headers, start=1):
        width = preferred_widths.get(header, 18)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _remove_temporary_file(path: object) -> None:
    if not isinstance(path, Path):
        return

    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
