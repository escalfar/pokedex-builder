import csv
import json
from pathlib import Path

from pokedex.constants import OUTPUT_COLUMNS, Gender
from pokedex.exporter_csv import export_csv
from pokedex.exporter_json import export_json
from pokedex.models import GameAvailability, PokemonEntry


def build_entry() -> PokemonEntry:
    return PokemonEntry(
        national_dex=26,
        pokemon="Raichu",
        form="Alolan",
        name="Alolan Raichu",
        generation=7,
        home_id="00026_ALOLA",
        availability=GameAvailability(),
        legendary_mythical=False,
        obtainable_shiny=True,
        gender=Gender.NONE,
    )


def test_export_csv_creates_expected_columns(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "Pokedex.csv"

    result = export_csv(
        (build_entry(),),
        output_path,
    )

    assert result == output_path
    assert output_path.is_file()

    with output_path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        rows = list(csv.DictReader(file))

    assert len(rows) == 1
    assert list(rows[0]) == [column.value for column in OUTPUT_COLUMNS]
    assert rows[0]["Nat Dex"] == "26"
    assert rows[0]["Nombre"] == "Alolan Raichu"
    assert rows[0]["Obtenido"] == "☐"
    assert rows[0]["Prioridad"] == ""
    assert rows[0]["XY"] == "FALSE"
    assert rows[0]["Obtenible"] == "TRUE"
    assert rows[0]["Posibles"] == ""


def test_export_json_preserves_native_booleans(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "Pokedex.json"

    result = export_json(
        (build_entry(),),
        output_path,
    )

    assert result == output_path

    payload = json.loads(output_path.read_text(encoding="utf-8"))

    assert len(payload) == 1
    assert payload[0]["Nat Dex"] == 26
    assert payload[0]["Nombre"] == "Alolan Raichu"
    assert payload[0]["Obtenido"] == "☐"
    assert payload[0]["Prioridad"] is None
    assert payload[0]["XY"] is False
    assert payload[0]["Obtenible"] is True
    assert payload[0]["Posibles"] is None


def test_json_uses_unescaped_unicode(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "Pokedex.json"

    export_json(
        (build_entry(),),
        output_path,
    )

    content = output_path.read_text(encoding="utf-8")

    assert "Legendario/Mítico" in content
    assert "\\u00ed" not in content


def test_export_excel_creates_four_expected_sheets(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    output_path = tmp_path / "Pokedex.xlsx"
    result = export_excel((build_entry(),), output_path)

    assert result == output_path
    workbook = load_workbook(output_path, data_only=True)

    try:
        assert workbook.sheetnames == [
            "Pokédex",
            "Resumen",
            "Validación",
            "Metadatos",
        ]
        sheet = workbook["Pokédex"]
        assert sheet.freeze_panes == "A2"
        assert sheet["A1"].value == "Nat Dex"
        assert sheet["F2"].value == "Raichu (Alolan)"
        assert sheet["G2"].value == "☐"
        assert sheet["H2"].value is None
        assert sheet["T2"].value is True
        assert sheet["U2"].value is None
        assert sheet.auto_filter.ref == sheet.dimensions
        assert len(sheet.tables) == 0
    finally:
        workbook.close()


def test_export_excel_uses_requested_column_order_and_hidden_columns(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    output_path = tmp_path / "Pokedex.xlsx"
    export_excel((build_entry(),), output_path)

    workbook = load_workbook(output_path, data_only=True)
    try:
        sheet = workbook["Pokédex"]
        headers = [cell.value for cell in sheet[1]]
        assert headers == [
            "Nat Dex",
            "Pokemon",
            "Forma",
            "Gen",
            "ID HOME",
            "Nombre",
            "Obtenido",
            "Prioridad",
            "XY",
            "ORAS",
            "SM",
            "USUM",
            "LGPE",
            "SwSh",
            "Arceus",
            "BDSP",
            "ScVi",
            "ZA",
            "Legendario/Mítico",
            "Obtenible",
            "Posibles",
        ]
        assert {
            column
            for column in ("B", "C", "D", "E", "S", "T", "U")
            if sheet.column_dimensions[column].hidden
        } == {"B", "C", "D", "E", "S", "T", "U"}
        assert sheet.column_dimensions["F"].hidden is False
        assert sheet.column_dimensions["G"].hidden is False
        assert sheet.column_dimensions["H"].hidden is False
    finally:
        workbook.close()


def test_export_excel_summary_and_metadata(tmp_path: Path) -> None:
    from datetime import UTC, datetime

    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    output_path = tmp_path / "Pokedex.xlsx"
    generated_at = datetime(2026, 7, 15, 12, 0, tzinfo=UTC)

    export_excel(
        (build_entry(),),
        output_path,
        generator_version="1.2.3",
        specification_version="1.0",
        generated_at=generated_at,
    )

    workbook = load_workbook(output_path, data_only=True)

    try:
        summary = workbook["Resumen"]
        metadata = workbook["Metadatos"]
        validation = workbook["Validación"]

        assert summary["B4"].value == 1
        assert summary["B5"].value == 1
        assert metadata["B4"].value == "1.2.3"
        assert metadata["B6"].value == generated_at.isoformat()
        assert validation["B4"].value == "OK"
    finally:
        workbook.close()


def test_export_excel_adds_possible_formulas_summary_and_dex_alignment(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    output_path = tmp_path / "Pokedex.xlsx"
    export_excel((build_entry(),), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        pokedex = workbook["Pokédex"]
        summary = workbook["Resumen"]

        assert pokedex["A2"].alignment.horizontal == "center"
        assert pokedex["U2"].value == '=OR(G2="☑",T2=TRUE)'
        assert pokedex.column_dimensions["U"].hidden is True

        assert summary["A7"].value == "Total obtenidos"
        assert summary["B7"].value == "=COUNTIF('Pokédex'!$G$2:$G$2,\"☑\")"
        assert summary["A8"].value == "Total no obtenidos"
        assert summary["B8"].value == "=COUNTIF('Pokédex'!$G$2:$G$2,\"☐\")"
        assert summary["A9"].value == "Total posibles"
        assert summary["B9"].value == "=COUNTIF('Pokédex'!$U$2:$U$2,TRUE)"
        assert summary["A10"].value == "Porcentaje obtenidos / posibles"
        assert summary["B10"].value == "=IFERROR(B7/B9,0)"
        assert summary["B10"].number_format == "0.00%"
    finally:
        workbook.close()


def test_excel_name_starts_with_species_without_changing_other_exports(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    entry = build_entry()
    output_path = tmp_path / "Pokedex.xlsx"
    export_excel((entry,), output_path)

    workbook = load_workbook(output_path, data_only=True)
    try:
        assert workbook["Pokédex"]["F2"].value == "Raichu (Alolan)"
    finally:
        workbook.close()

    assert entry.name == "Alolan Raichu"


def _build_form_entry(
    *,
    national_dex: int,
    pokemon: str,
    form: str,
    home_id: str,
) -> PokemonEntry:
    return PokemonEntry(
        national_dex=national_dex,
        pokemon=pokemon,
        form=form,
        name=pokemon if form == "Normal" else f"{form} {pokemon}",
        generation=6,
        home_id=home_id,
        availability=GameAvailability(),
        legendary_mythical=False,
        obtainable_shiny=True,
        gender=Gender.NONE,
    )


def test_excel_adds_selected_base_form_labels_only_to_nombre(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    expected_labels = {
        "Deoxys": "Normal",
        "Wormadam": "Plant",
        "Shaymin": "Land",
        "Basculin": "Red",
        "Pumpkaboo": "Medium",
        "Gourgeist": "Medium",
        "Oricorio": "Baile",
        "Lycanroc": "Midday",
        "Toxtricity": "Amped",
        "Urshifu": "Single",
        "Squawkabilly": "Green",
        "Tatsugiri": "Curly",
        "Gimmighoul": "Chest",
    }
    entries = tuple(
        _build_form_entry(
            national_dex=index,
            pokemon=pokemon,
            form="Normal",
            home_id=f"{index:05d}_NORMAL_NONE",
        )
        for index, pokemon in enumerate(expected_labels, start=1)
    )
    output_path = tmp_path / "Pokedex.xlsx"
    export_excel(entries, output_path)

    workbook = load_workbook(output_path, data_only=True)
    try:
        names = [
            workbook["Pokédex"].cell(row=row, column=6).value
            for row in range(2, len(entries) + 2)
        ]
    finally:
        workbook.close()

    assert names == [
        f"{pokemon} ({label})" for pokemon, label in expected_labels.items()
    ]
    assert [entry.name for entry in entries] == list(expected_labels)


def test_excel_uses_matching_flower_order_and_eternal_floette_last(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    forms_by_species = {
        "Flabébé": (
            "Blue Flower",
            "Red Flower",
            "White Flower",
            "Orange Flower",
            "Yellow Flower",
        ),
        "Floette": (
            "Eternal",
            "Blue Flower",
            "Red Flower",
            "White Flower",
            "Orange Flower",
            "Yellow Flower",
        ),
        "Florges": (
            "Blue Flower",
            "Red Flower",
            "White Flower",
            "Orange Flower",
            "Yellow Flower",
        ),
    }
    dex_numbers = {"Flabébé": 669, "Floette": 670, "Florges": 671}
    entries = tuple(
        _build_form_entry(
            national_dex=dex_numbers[pokemon],
            pokemon=pokemon,
            form=form,
            home_id=f"{dex_numbers[pokemon]:05d}_{form.upper().replace(' ', '_')}_NONE",
        )
        for pokemon, forms in forms_by_species.items()
        for form in forms
    )
    output_path = tmp_path / "Pokedex.xlsx"
    export_excel(entries, output_path)

    workbook = load_workbook(output_path, data_only=True)
    try:
        names = [
            workbook["Pokédex"].cell(row=row, column=6).value
            for row in range(2, len(entries) + 2)
        ]
    finally:
        workbook.close()

    common_order = (
        "Red Flower",
        "Yellow Flower",
        "Orange Flower",
        "Blue Flower",
        "White Flower",
    )
    expected_colors = ("Red", "Yellow", "Orange", "Blue", "White")
    assert names == [
        *(f"Flabébé ({color})" for color in expected_colors),
        *(f"Floette ({color})" for color in expected_colors),
        "Floette (Eternal)",
        *(f"Florges ({color})" for color in expected_colors),
    ]


def test_excel_simplifies_selected_form_labels_only_in_nombre(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    cases = (
        (128, "Tauros", "Paldean Aqua Breed", "Tauros (Paldean Aqua)"),
        (128, "Tauros", "Paldean Blaze Breed", "Tauros (Paldean Blaze)"),
        (128, "Tauros", "Paldean Combat Breed", "Tauros (Paldean Combat)"),
        (412, "Burmy", "Sandy Cloak", "Burmy (Sandy)"),
        (412, "Burmy", "Trash Cloak", "Burmy (Trash)"),
        (550, "Basculin", "Blue Striped", "Basculin (Blue)"),
        (550, "Basculin", "White Striped", "Basculin (White)"),
        (666, "Vivillon", "Poké Ball Pattern", "Vivillon (Poké Ball)"),
        (710, "Pumpkaboo", "Super", "Pumpkaboo (Jumbo)"),
        (711, "Gourgeist", "Super", "Gourgeist (Jumbo)"),
        (892, "Urshifu", "Rapid Strike", "Urshifu (Rapid)"),
        (931, "Squawkabilly", "Blue Plumage", "Squawkabilly (Blue)"),
        (931, "Squawkabilly", "White Plumage", "Squawkabilly (White)"),
        (931, "Squawkabilly", "Yellow Plumage", "Squawkabilly (Yellow)"),
    )
    entries = tuple(
        _build_form_entry(
            national_dex=dex,
            pokemon=pokemon,
            form=form,
            home_id=f"{dex:05d}_{index}_NONE",
        )
        for index, (dex, pokemon, form, _) in enumerate(cases, start=1)
    )
    output_path = tmp_path / "Pokedex.xlsx"
    export_excel(entries, output_path)

    workbook = load_workbook(output_path, data_only=True)
    try:
        names = [
            workbook["Pokédex"].cell(row=row, column=6).value
            for row in range(2, len(entries) + 2)
        ]
    finally:
        workbook.close()

    assert names == [expected for *_, expected in cases]
    assert [entry.form for entry in entries] == [form for _, _, form, _ in cases]


def test_export_excel_adds_tracking_validations_and_conditional_formats(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    from pokedex.exporter_excel import export_excel

    output_path = tmp_path / "Pokedex.xlsx"
    export_excel((build_entry(),), output_path)

    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Pokédex"]
        validations = list(sheet.data_validations.dataValidation)

        assert len(validations) == 2
        obtained_validation = next(
            validation for validation in validations if validation.type == "list"
        )
        priority_validation = next(
            validation for validation in validations if validation.type == "whole"
        )

        assert obtained_validation.formula1 == '"☐,☑"'
        assert obtained_validation.allow_blank is False
        assert str(obtained_validation.sqref) == "G2"

        assert priority_validation.operator == "between"
        assert priority_validation.formula1 == "0"
        assert priority_validation.formula2 == "10"
        assert priority_validation.allow_blank is True
        assert str(priority_validation.sqref) == "H2"

        conditional_ranges = {str(key) for key in sheet.conditional_formatting}
        assert conditional_ranges == {
            "<ConditionalFormatting G2>",
            "<ConditionalFormatting H2>",
        }

        obtained_rules = list(sheet.conditional_formatting["G2"])
        priority_rules = list(sheet.conditional_formatting["H2"])
        assert len(obtained_rules) == 1
        assert obtained_rules[0].type == "cellIs"
        assert obtained_rules[0].operator == "equal"
        assert obtained_rules[0].formula == ['"☑"']
        assert obtained_rules[0].stopIfTrue is True

        assert len(priority_rules) == 11
        assert all(rule.type == "cellIs" for rule in priority_rules)
        assert [rule.formula for rule in priority_rules] == [
            [str(value)] for value in range(11)
        ]
        assert all(rule.stopIfTrue is True for rule in priority_rules)
    finally:
        workbook.close()
